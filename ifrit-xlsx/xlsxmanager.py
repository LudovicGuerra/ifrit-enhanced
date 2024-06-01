import copy
import os
import re
from math import floor

import xlsxwriter
from openpyxl.reader.excel import load_workbook

from ennemy import Ennemy
from gamedata import GameData

COL_MONSTER_INFO = 0
COL_STAT = 3
COL_DEF = 14
COL_ITEM = 17
COL_MISC = 25
COL_ABILITIES = 28
ROW_FILE_DATA = 42
COL_FILE_DATA = 0
ROW_MAG = 1
ROW_MUG = 5
ROW_DROP = 9
ROW_MONSTER_NAME = 1
ROW_MONSTER_LVL = 2
ROW_MONSTER_NB_ANIMATION = 3
ROW_MONSTER_COMBAT_TEXT = 4

ROW_BYTE_FLAG = 8
ROW_RENZOKUKEN = 41

COL_SHEET_LOW_LVL = 1
COL_SHEET_MED_LVL = 3
COL_SHEET_HIGH_LVL = 5
DEFAULT_MONSTER_LVL = 10
COL_GRAPH_PER_LVL = 4
ROW_GRAPH_PER_LVL = 33
NB_MAX_ABILITIES = 16

STAT_GRAPH_CELL_PLACEMENT = 'N34'
STAT_GRAPH_WIDTH = 850
STAT_GRAPH_HEIGHT = 500

ROW_DROP_CARD = 20
COL_DROP_CARD = 28
ROW_DEVOUR = 24
COL_DEVOUR = 28

REF_DATA_COL_ABILITIES_TYPE = 0
REF_DATA_COL_ABILITIES = 1
REF_DATA_COL_MAGIC = 2
REF_DATA_COL_ITEM = 3
REF_DATA_COL_CARD = 4
REF_DATA_COL_DEVOUR = 5
REF_DATA_COL_SPECIAL_ACTION = 6
REF_DATA_COL_LIST = [REF_DATA_COL_ABILITIES_TYPE, REF_DATA_COL_ABILITIES, REF_DATA_COL_MAGIC, REF_DATA_COL_ITEM, REF_DATA_COL_CARD, REF_DATA_COL_DEVOUR,
                     REF_DATA_COL_SPECIAL_ACTION]
REF_DATA_SHEET_TITLE = 'ref_data'

MAX_COMBAT_TXT = 8
MAX_SHEET_TITLE_SIZE = 31
INVALID_CHAR_TITLE_EXCEL_LIST = ['[', ']', ':', '*', '?', '/', '\\']


class DatToXlsx():

    def __init__(self, ifrit_xlsx):
        self.workbook = xlsxwriter.Workbook(ifrit_xlsx)
        self.__init_style()
        self.__file_name_list = []

    def __init_style(self):
        self.column_title_style = self.workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#b2b2b2', 'align': 'center'})
        self.border_style = self.workbook.add_format({'border': 1})
        self.row_title_style = self.workbook.add_format({'border': 1, 'bold': True})
        self.magic_style = self.workbook.add_format({'border': 1, 'bg_color': '#b8cce4'})
        self.status_style = self.workbook.add_format({'border': 1, 'bg_color': '#b9b085'})
        self.drop_style = self.workbook.add_format({'border': 1, 'bg_color': '#ccc0da'})
        self.mug_style = self.workbook.add_format({'border': 1, 'bg_color': '#7aeca0'})
        self.magic_style = self.workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#b8cce4'})
        self.status_style = self.workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#b9b085'})
        self.drop_style = self.workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#ccc0da'})
        self.mug_style = self.workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#7aeca0'})
        self.percent_style = self.workbook.add_format({'num_format': '#,##0.00%', 'border': 1})

    def __hp_excel_formula(self, stat_cell: list, monster_lvl: int):
        return ('=(FLOOR({}*({}*{}/20+{}),1) + 10*{} + {}*100*{} + 1000*{})/100'
                .format(stat_cell[0], monster_lvl, monster_lvl, monster_lvl, stat_cell[1], stat_cell[2], monster_lvl, stat_cell[3]))

    def __str_excel_formula(self, stat_cell: list, monster_lvl: int):
        return ('=FLOOR({}*{}/40, 1)+FLOOR({}/(4*{}),1)+FLOOR({}/4,1)+FLOOR({}*{}/(8*{}),1)'
                .format(monster_lvl, stat_cell[0], monster_lvl, stat_cell[1], stat_cell[2], monster_lvl, monster_lvl, stat_cell[3]))

    def __create_title(self, game_data, worksheet):
        worksheet.write_row(0, COL_MONSTER_INFO, ["Monster info"], cell_format=self.column_title_style)
        worksheet.write_row(0, COL_STAT + 1,
                            ["Value 1", "Value 2", "Value 3", "Value 4", "Impact 1", "Impact 2", "Impact 3", "Impact 4", "Total"],
                            cell_format=self.column_title_style)
        worksheet.write_row(0, COL_DEF, ["Defense name", "% Resistance"], cell_format=self.column_title_style)
        worksheet.write_row(0, COL_ITEM + 1, ["Low Level", "Number", "Medium Level", "Number", "High Level", "Number"],
                            cell_format=self.column_title_style)
        worksheet.write_row(0, COL_MISC, ["Property name", "Value"], cell_format=self.column_title_style)
        worksheet.merge_range(xlsxwriter.utility.xl_col_to_name(COL_ABILITIES) + "1:" + xlsxwriter.utility.xl_col_to_name(COL_ABILITIES + 2) + "1",
                              "Low Level",
                              cell_format=self.column_title_style)
        worksheet.merge_range(xlsxwriter.utility.xl_col_to_name(COL_ABILITIES + 3) + "1:" + xlsxwriter.utility.xl_col_to_name(COL_ABILITIES + 5) + "1",
                              "Medium Level", cell_format=self.column_title_style)
        worksheet.merge_range(xlsxwriter.utility.xl_col_to_name(COL_ABILITIES + 6) + "1:" + xlsxwriter.utility.xl_col_to_name(COL_ABILITIES + 8) + "1",
                              "High Level", cell_format=self.column_title_style)
        worksheet.write_row(1, COL_ABILITIES, ["Type", "Ability", "Animation", "Type", "Ability", "Animation", "Type", "Ability", "Animation"],
                            cell_format=self.column_title_style)
        worksheet.write(ROW_DROP_CARD, COL_DROP_CARD, game_data.CARD_DATA['pretty_name'], self.column_title_style)
        worksheet.write(ROW_DEVOUR, COL_DEVOUR, game_data.DEVOUR_DATA['pretty_name'], self.column_title_style)
        worksheet.write(ROW_RENZOKUKEN, COL_MISC, game_data.SECTION_INFO_STAT_RENZOKUKEN['pretty_name'], self.column_title_style)

    def __get_file_index(self, ennemy):
        return int(re.search(r'\d{3}', ennemy.origin_file_name).group())

    def __get_tab_name(self, ennemy):
        file_name = str(self.__get_file_index(ennemy)) + " - " + ennemy.info_stat_data['monster_name']
        if file_name == '':
            file_name = "Empty"
        while file_name in self.__file_name_list:
            file_name += " dub"

        if len(file_name) > MAX_SHEET_TITLE_SIZE:
            file_name = file_name[:MAX_SHEET_TITLE_SIZE]
        for char in INVALID_CHAR_TITLE_EXCEL_LIST:
            file_name = file_name.replace(char, ';')
        self.__file_name_list.append(file_name)
        return file_name

    def __validate_elem_def(self, worksheet, game_data, row_index, column_index):
        worksheet.data_validation(row_index, column_index, row_index, column_index,
                                  {'validate': 'integer', 'criteria': 'between',
                                   'minimum': game_data.ELEM_DEF_MIN_VAL, 'maximum': game_data.ELEM_DEF_MAX_VAL,
                                   'input_title': 'Elem def',
                                   'input_message': 'Between ' + str(game_data.ELEM_DEF_MIN_VAL) + ' and ' + str(game_data.ELEM_DEF_MAX_VAL)})

    def __validate_status_def(self, worksheet, game_data, row_index, column_index):
        worksheet.data_validation(row_index, column_index, row_index, column_index,
                                  {'validate': 'integer', 'criteria': 'between',
                                   'minimum': game_data.STATUS_DEF_MIN_VAL, 'maximum': game_data.STATUS_DEF_MAX_VAL,
                                   'input_title': 'Status def',
                                   'input_message': 'Between ' + str(game_data.STATUS_DEF_MIN_VAL) + ' and ' + str(game_data.STATUS_DEF_MAX_VAL)})

    def __validation_post_process_all(self, worksheet, game_data: GameData):
        self.__validate_stat(worksheet, game_data)
        self.__validate_card(worksheet, game_data)
        self.__validate_devour(worksheet, game_data)
        self.__validate_draw_drop_mug(worksheet, game_data)
        self.__validate_renzokuken(worksheet, game_data)

    def __validate_stat(self, worksheet, game_data: GameData):
        worksheet.data_validation(1, COL_STAT, 1 + 7, COL_STAT + 4,
                                  {'validate': 'integer', 'criteria': 'between',
                                   'minimum': game_data.STAT_MIN_VAL, 'maximum': game_data.STAT_MAX_VAL,
                                   'input_title': 'Stat',
                                   'input_message': 'Between ' + str(game_data.STAT_MIN_VAL) + ' and ' + str(game_data.STAT_MAX_VAL)})

    def __validate_card(self, worksheet, game_data: GameData):
        col_str = xlsxwriter.utility.xl_col_to_name(REF_DATA_COL_CARD)
        source_str = '=' + REF_DATA_SHEET_TITLE + '!$' + col_str + '2:$' + col_str + '$' + str(len(game_data.card_values) + 1)
        worksheet.data_validation(ROW_DROP_CARD + 1, COL_DROP_CARD + 1, ROW_DROP_CARD + 1 + 2, COL_DROP_CARD + 1,
                                  {'validate': 'list', 'source': source_str})

    def __validate_devour(self, worksheet, game_data: GameData):
        col_str = xlsxwriter.utility.xl_col_to_name(REF_DATA_COL_DEVOUR)
        source_str = '=' + REF_DATA_SHEET_TITLE + '!$' + col_str + '2:$' + col_str + '$' + str(len(game_data.devour_values) + 1)
        worksheet.data_validation(ROW_DEVOUR + 1, COL_DEVOUR + 1, ROW_DEVOUR + 1 + 2, COL_DEVOUR + 1,
                                  {'validate': 'list', 'source': source_str})

    def __validate_renzokuken(self, worksheet, game_data: GameData):
        col_str = xlsxwriter.utility.xl_col_to_name(REF_DATA_COL_SPECIAL_ACTION)
        source_str = '=' + REF_DATA_SHEET_TITLE + '!$' + col_str + '2:$' + col_str + '$' + str(len(game_data.special_action) + 1)
        worksheet.data_validation(ROW_RENZOKUKEN + 1, COL_MISC + 1, ROW_RENZOKUKEN + 1 + 2, COL_MISC + 1,
                                  {'validate': 'list', 'source': source_str})

    def __validate_draw_drop_mug(self, worksheet, game_data: GameData):
        col_values = [COL_SHEET_LOW_LVL, COL_SHEET_MED_LVL, COL_SHEET_HIGH_LVL]
        col_str = xlsxwriter.utility.xl_col_to_name(REF_DATA_COL_MAGIC)
        source_str = '=' + REF_DATA_SHEET_TITLE + '!$' + col_str + '2:$' + col_str + '$' + str(len(game_data.magic_values) + 1)
        #ROW_MAG, COL_ABILITIES + col_value, ROW_MAG + 3, COL_ABILITIES + col_value,
        for col_value in col_values:
            worksheet.data_validation(ROW_MAG, COL_ITEM + col_value, ROW_MAG + 3, COL_ITEM + col_value, {'validate': 'list', 'source': source_str})
        col_str = xlsxwriter.utility.xl_col_to_name(REF_DATA_COL_ITEM)
        source_str = '=' + REF_DATA_SHEET_TITLE + '!$' + col_str + '2:$' + col_str + '$' + str(len(game_data.item_values) + 1)
        row_values = [ROW_MUG, ROW_DROP]
        for col_value in col_values:
            for row_value in row_values:
                worksheet.data_validation(row_value, COL_ITEM + col_value, row_value + 3, COL_ITEM + col_value,
                                          {'validate': 'list', 'source': source_str})

    def __validate_abilities(self, worksheet, game_data, ability_type, id, row_index, column_index):
        if ability_type['name'] == "Magic":
            ability_name = game_data.magic_values[id]['ref']
            col_ab_str = xlsxwriter.utility.xl_col_to_name(REF_DATA_COL_MAGIC)
        elif ability_type['name'] == "Custom":
            ability_name = game_data.ennemy_abilities_values[id]['ref']
            col_ab_str = xlsxwriter.utility.xl_col_to_name(REF_DATA_COL_ABILITIES)
        elif ability_type['name'] == "Item":
            ability_name = game_data.item_values[id]['ref']
            col_ab_str = xlsxwriter.utility.xl_col_to_name(REF_DATA_COL_ITEM)
        else:
            if id < len(game_data.ennemy_abilities_values):
                ability_name = game_data.ennemy_abilities_values[id]['ref']
            else:
                game_data.ennemy_abilities_values[id] = {'name': "Temp Garbage", 'ref': str(id) + ":Temp Garbage"}
                ability_name = game_data.ennemy_abilities_values[id]['ref']
            col_ab_str = xlsxwriter.utility.xl_col_to_name(REF_DATA_COL_ABILITIES)

        col_str = xlsxwriter.utility.xl_col_to_name(REF_DATA_COL_ABILITIES_TYPE)
        source_str = '=' + REF_DATA_SHEET_TITLE + '!$' + col_str + '2:$' + col_str + '$' + str(len(game_data.ennemy_abilities_type_values) + 1)
        worksheet.data_validation(row_index, column_index, row_index, column_index,
                                  {'validate': 'list', 'source': source_str})
        worksheet.write(row_index, column_index + 1, ability_name, self.border_style)
        source_str = '=' + REF_DATA_SHEET_TITLE + '!$' + col_ab_str + '2:$' + col_ab_str + '$' + str(len(game_data.ennemy_abilities_values) + 1)
        worksheet.data_validation(row_index, column_index + 1, row_index,
                                  column_index + 1,
                                  {'validate': 'list', 'source': source_str})

    def export_to_xlsx(self, ennemy_list: dict, game_data):
        # Chart
        ## Stat chart
        chart_stat = {}
        tab_list = []
        for key, ennemy in ennemy_list.items():
            # Tab (sheet) creation. Checking name doesn't already exist
            file_index = self.__get_file_index(ennemy)
            if file_index == 127 or file_index > 143:  # Ignoring garbage files
                continue

            chart_stat[ennemy] = self.workbook.add_chart({'type': 'line'})
            file_name = self.__get_tab_name(ennemy)
            worksheet = self.workbook.add_worksheet(file_name)
            print("Write to XLSX file {}".format(file_name))

            # Column position of different "menu"
            column_index = {}
            column_index['stat'] = COL_STAT
            column_index['def'] = COL_DEF
            column_index['item'] = COL_ITEM
            column_index['misc'] = COL_MISC
            column_index['graph_stat'] = COL_GRAPH_PER_LVL + 1
            column_index['abilities'] = COL_ABILITIES

            # Titles
            self.__create_title(game_data, worksheet)

            # File info not link to the monster data
            worksheet.write(ROW_FILE_DATA, COL_FILE_DATA, "File data", self.column_title_style)
            worksheet.write(ROW_FILE_DATA + 1, COL_FILE_DATA, "Original file name", self.row_title_style)
            worksheet.write(ROW_FILE_DATA + 1, COL_FILE_DATA + 1, key, self.border_style)

            # Graph level
            worksheet.write(ROW_GRAPH_PER_LVL, COL_GRAPH_PER_LVL, "Level", self.column_title_style)
            for i in range(1, 101):
                worksheet.write(ROW_GRAPH_PER_LVL + i, COL_GRAPH_PER_LVL, i, self.row_title_style)

            # Index setting
            row_index = {}
            row_index['stat'] = 1
            row_index['stat'] = 1
            row_index['def'] = 1
            row_index['item'] = 1
            row_index['misc'] = 1
            row_index['abilities'] = 2
            row_index['byte_flag'] = 0
            index = {}
            index['elem_def'] = 0
            index['status_def'] = 0
            # General Monster info
            worksheet.write(ROW_MONSTER_LVL, COL_MONSTER_INFO, "Monster LVL", self.row_title_style)
            worksheet.write(ROW_MONSTER_LVL, COL_MONSTER_INFO + 1, DEFAULT_MONSTER_LVL, self.border_style)
            worksheet.write(ROW_MONSTER_NAME, COL_MONSTER_INFO, "Name", self.row_title_style)
            worksheet.write(ROW_MONSTER_NAME, COL_MONSTER_INFO + 1, ennemy.info_stat_data['monster_name'], self.border_style)
            worksheet.write(ROW_MONSTER_NB_ANIMATION, COL_MONSTER_INFO, "Nb animation", self.row_title_style)
            worksheet.write(ROW_MONSTER_NB_ANIMATION, COL_MONSTER_INFO + 1, ennemy.model_animation_data['nb_animation'], self.border_style)

            for i in range(len(ennemy.battle_script_data['battle_text'])):
                worksheet.write(ROW_MONSTER_COMBAT_TEXT + i, COL_MONSTER_INFO, "Combat text {}".format(i), self.row_title_style)
                worksheet.write(ROW_MONSTER_COMBAT_TEXT + i, COL_MONSTER_INFO + 1, "{}".format(ennemy.battle_script_data['battle_text'][i]), self.border_style)

            # Filling the Excel
            for param_name, value in ennemy.info_stat_data.items():
                # Search the pretty_name
                pretty_name = ""
                for data_el in game_data.SECTION_INFO_STAT_LIST_DATA:
                    if data_el['name'] == param_name:
                        pretty_name = data_el['pretty_name']
                        break
                try:
                    # Stat menu
                    column_index['stat'] = COL_STAT
                    impact = []
                    if param_name in game_data.stat_values:
                        # Writing title of row
                        worksheet.write(row_index['stat'], COL_STAT, pretty_name, self.row_title_style)
                        column_index['stat'] += 1
                        # Writing the 4 values of stats
                        for el2 in value:
                            worksheet.write(row_index['stat'], column_index['stat'], el2, self.border_style)
                            column_index['stat'] += 1

                        # Writing the Impact and total stat
                        monster_lvl_cell = xlsxwriter.utility.xl_col_to_name(COL_MONSTER_INFO + 1) + str(ROW_MONSTER_LVL + 1)
                        stat_cell = [None] * 4
                        # Title lvl column stat
                        worksheet.write(ROW_GRAPH_PER_LVL, column_index['graph_stat'], pretty_name, self.column_title_style)
                        if param_name == 'hp':
                            # Impact 1
                            stat_cell[0] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 1) + str(row_index['stat'] + 1)
                            worksheet.write(row_index['stat'], column_index['stat'],
                                            '=FLOOR({}*({}*{}/20+{}),1)'.format(stat_cell[0], monster_lvl_cell, monster_lvl_cell, monster_lvl_cell),
                                            self.border_style)
                            # Impact 2
                            stat_cell[1] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 2) + str(row_index['stat'] + 1)
                            worksheet.write(row_index['stat'], column_index['stat'] + 1, '=10*{}'.format(stat_cell[1]), self.border_style)
                            # Impact 3
                            stat_cell[2] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 3) + str(row_index['stat'] + 1)
                            worksheet.write(row_index['stat'], column_index['stat'] + 2, '={}*100*{}'.format(stat_cell[2], monster_lvl_cell), self.border_style)
                            # Impact 4
                            stat_cell[3] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 4) + str(row_index['stat'] + 1)
                            worksheet.write(row_index['stat'], column_index['stat'] + 3, '=1000*{}'.format(stat_cell[3]), self.border_style)
                            # Total Sum
                            worksheet.write(row_index['stat'], column_index['stat'] + 4,
                                            '=SUM({}:{})'.format(xlsxwriter.utility.xl_col_to_name(COL_STAT + 5) + str(row_index['stat'] + 1),
                                                                 xlsxwriter.utility.xl_col_to_name(COL_STAT + 8) + str(row_index['stat'] + 1)),
                                            self.border_style)
                            # Total Formula
                            # The HP is harder to show on a graph, so we divide it by 100 to show it better
                            worksheet.write(ROW_GRAPH_PER_LVL, column_index['graph_stat'], pretty_name, self.column_title_style)
                            for i in range(1, 101):
                                monster_lvl_cell = xlsxwriter.utility.xl_col_to_name(COL_GRAPH_PER_LVL) + str(ROW_GRAPH_PER_LVL + i + 1)
                                # i corresponding to the monster level
                                # /101 because HP is too high to be plot correctly with others values.
                                str_formula = self.__hp_excel_formula(stat_cell, monster_lvl_cell)
                                worksheet.write(ROW_GRAPH_PER_LVL + i, column_index['graph_stat'], str_formula, self.border_style)

                        elif param_name == 'str' or param_name == 'mag':
                            # Impact 1
                            stat_cell[0] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 1) + str(row_index['stat'] + 1)
                            worksheet.write(row_index['stat'], column_index['stat'], '=FLOOR({}*{}/40, 1)'.format(monster_lvl_cell, stat_cell[0]),
                                            self.border_style)
                            # Impact 2
                            stat_cell[1] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 2) + str(row_index['stat'] + 1)
                            worksheet.write(row_index['stat'], column_index['stat'] + 1, '=FLOOR({}/(4*{}),1)'.format(monster_lvl_cell, stat_cell[1]),
                                            self.border_style)
                            # Impact 3
                            stat_cell[2] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 3) + str(row_index['stat'] + 1)
                            worksheet.write(row_index['stat'], column_index['stat'] + 2, '=FLOOR({}/4,1)'.format(stat_cell[2]), self.border_style)
                            # Impact 4
                            stat_cell[3] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 4) + str(row_index['stat'] + 1)
                            worksheet.write(row_index['stat'], column_index['stat'] + 3,
                                            '=FLOOR({}*{}/(8*{}),1)'.format(monster_lvl_cell, monster_lvl_cell, stat_cell[3]), self.border_style)
                            # Total
                            worksheet.write(row_index['stat'], column_index['stat'] + 4,
                                            '=SUM({}:{})'.format(xlsxwriter.utility.xl_col_to_name(COL_STAT + 5) + str(row_index['stat'] + 1),
                                                                 xlsxwriter.utility.xl_col_to_name(COL_STAT + 8) + str(row_index['stat'] + 1)),
                                            self.border_style)
                            # Total Formula
                            for i in range(1, 101):
                                monster_lvl_cell = xlsxwriter.utility.xl_col_to_name(COL_GRAPH_PER_LVL) + str(ROW_GRAPH_PER_LVL + i + 1)
                                # i corresponding to the monster level
                                str_formula = self.__str_excel_formula(stat_cell, monster_lvl_cell)
                                worksheet.write(ROW_GRAPH_PER_LVL + i, column_index['graph_stat'], str_formula, self.border_style)
                        elif param_name == 'vit' or param_name == 'spr' or param_name == 'spd' or param_name == 'eva':
                            # Impact 1
                            stat_cell[0] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 1) + str(row_index['stat'] + 1)
                            worksheet.write(row_index['stat'], column_index['stat'], '={}*{}'.format(monster_lvl_cell, stat_cell[0]), self.border_style)
                            # Impact 2
                            stat_cell[1] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 2) + str(row_index['stat'] + 1)
                            worksheet.write(row_index['stat'], column_index['stat'] + 1, '=FLOOR({}/{},1)'.format(monster_lvl_cell, stat_cell[1]),
                                            self.border_style)
                            # Impact 3
                            stat_cell[2] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 3) + str(row_index['stat'] + 1)
                            worksheet.write(row_index['stat'], column_index['stat'] + 2, '={}'.format(stat_cell[2]), self.border_style)
                            # Impact 4
                            stat_cell[3] = xlsxwriter.utility.xl_col_to_name(COL_STAT + 4) + str(row_index['stat'] + 1)
                            worksheet.write(row_index['stat'], column_index['stat'] + 3, '=-FLOOR({}/{},1)'.format(monster_lvl_cell, stat_cell[3]),
                                            self.border_style)
                            # Total
                            worksheet.write(row_index['stat'], column_index['stat'] + 4,
                                            '=SUM({}:{})'.format(xlsxwriter.utility.xl_col_to_name(COL_STAT + 5) + str(row_index['stat'] + 1),
                                                                 xlsxwriter.utility.xl_col_to_name(COL_STAT + 8) + str(row_index['stat'] + 1)),
                                            self.border_style)
                            # Total Formula
                            for i in range(1, 101):
                                monster_lvl_cell = xlsxwriter.utility.xl_col_to_name(COL_GRAPH_PER_LVL) + str(ROW_GRAPH_PER_LVL + i + 1)
                                # i corresponding to the monster level
                                str_formula = '={}*{}+FLOOR({}/{},1)+{}-FLOOR({}/{},1)'.format(monster_lvl_cell, stat_cell[0], monster_lvl_cell, stat_cell[1],
                                                                                               stat_cell[2], monster_lvl_cell, stat_cell[3])
                                worksheet.write(ROW_GRAPH_PER_LVL + i, column_index['graph_stat'], str_formula, self.border_style)

                        row_index['stat'] += 1
                        column_index['graph_stat'] += 1

                        # Creating chart with data computed
                        current_stat_column_str = xlsxwriter.utility.xl_col_to_name(column_index['graph_stat'] - 1)
                        lvl_range_str = '=\'' + file_name + '\'!$E$35:$E$135'
                        lvl_stat_str = '=\'' + file_name + '\'!${}$35:${}$135'.format(current_stat_column_str, current_stat_column_str)
                        graph_serie_name = pretty_name
                        if param_name == 'hp':
                            graph_serie_name += '/100'
                        chart_stat[ennemy].add_series({'name': graph_serie_name, 'categories': lvl_range_str, 'values': lvl_stat_str, 'smooth': True})
                    # Def menu
                    elif param_name in ['elem_def', 'status_def']:
                        for el2 in value:
                            if param_name == 'elem_def':
                                worksheet.write(row_index['def'], column_index['def'], game_data.magic_type_values[index[param_name]], self.magic_style)
                                worksheet.write(row_index['def'], column_index['def'] + 1, el2, self.magic_style)
                                self.__validate_elem_def(worksheet, game_data, row_index['def'], column_index['def'] + 1)
                            elif param_name == 'status_def':
                                worksheet.write(row_index['def'], column_index['def'], game_data.status_values[index[param_name]], self.status_style)
                                worksheet.write(row_index['def'], column_index['def'] + 1, el2, self.status_style)
                                self.__validate_status_def(worksheet, game_data, row_index['def'], column_index['def'] + 1)
                            index[param_name] += 1
                            row_index['def'] += 1
                    # Item menu (containing draw too)
                    elif param_name in ['low_lvl_mag', 'med_lvl_mag', 'high_lvl_mag', 'low_lvl_mug', 'med_lvl_mug', 'high_lvl_mug', 'low_lvl_drop',
                                        'med_lvl_drop',
                                        'high_lvl_drop']:  # Items
                        if 'mag' in param_name:
                            row_index['item'] = ROW_MAG
                        elif 'mug' in param_name:
                            row_index['item'] = ROW_MUG
                        elif 'drop' in param_name:
                            row_index['item'] = ROW_DROP
                        # Index as we fill the data first per 'type' (draw, mug, drop) then per column. And there is 3 column.
                        index['draw'] = 1
                        index['mug'] = 1
                        index['drop'] = 1
                        # Going through value
                        for el2 in value:
                            if 'low' in param_name:
                                col_index = column_index['item'] + COL_SHEET_LOW_LVL
                            elif 'med' in param_name:
                                col_index = column_index['item'] + COL_SHEET_MED_LVL
                            elif 'high' in param_name:
                                col_index = column_index['item'] + COL_SHEET_HIGH_LVL
                            else:  # Should not happen
                                print("Problem on column index")
                                col_index = column_index['item'] + COL_SHEET_HIGH_LVL + 2
                            if 'mag' in param_name:
                                worksheet.write(row_index['item'], column_index['item'], "Draw {}".format(index['draw']), self.magic_style)
                                index['draw'] += 1
                                worksheet.write(row_index['item'], col_index, game_data.magic_values[el2['ID']]['ref'], self.magic_style)
                                worksheet.write(row_index['item'], col_index + 1, el2['value'], self.magic_style)
                            elif 'mug' in param_name:
                                worksheet.write(row_index['item'], column_index['item'], "Mug {}".format(index['mug']), self.mug_style)
                                index['mug'] += 1
                                worksheet.write(row_index['item'], col_index, game_data.item_values[el2['ID']]['ref'], self.mug_style)
                                worksheet.write(row_index['item'], col_index + 1, el2['value'], self.mug_style)
                            elif 'drop' in param_name:
                                worksheet.write(row_index['item'], column_index['item'], "Drop {}".format(index['drop']), self.drop_style)
                                index['drop'] += 1
                                worksheet.write(row_index['item'], col_index, game_data.item_values[el2['ID']]['ref'], self.drop_style)
                                worksheet.write(row_index['item'], col_index + 1, el2['value'], self.drop_style)
                            row_index['item'] += 1
                    # Misc menu
                    elif param_name in game_data.MISC_ORDER:
                        worksheet.write(row_index['misc'], column_index['misc'], pretty_name, self.row_title_style)
                        if param_name == "mug_rate" or param_name == "drop_rate":  # Percent style need to divide by 100
                            worksheet.write(row_index['misc'], column_index['misc'] + 1, value / 100, self.percent_style)
                        else:
                            worksheet.write(row_index['misc'], column_index['misc'] + 1, floor(value), self.border_style)
                        row_index['misc'] += 1
                    # Abilities menu
                    elif param_name in game_data.ABILITIES_HIGHNESS_ORDER:
                        for el2 in value:
                            ability_type = game_data.ennemy_abilities_type_values[el2['type']]
                            worksheet.write(row_index['abilities'], column_index['abilities'], ability_type['ref'], self.border_style)
                            # Excel data validation
                            self.__validate_abilities(worksheet, game_data, ability_type, el2['id'], row_index['abilities'], column_index['abilities'])

                            worksheet.write(row_index['abilities'], column_index['abilities'] + 2, el2['animation'], self.border_style)
                            row_index['abilities'] += 1
                        row_index['abilities'] = 2
                        column_index['abilities'] += 3
                    elif param_name in ['card']:
                        worksheet.write(ROW_DROP_CARD + 1, COL_DROP_CARD, 'Drop', self.row_title_style)
                        worksheet.write(ROW_DROP_CARD + 1, COL_DROP_CARD + 1, game_data.card_values[value[0]]['ref'], self.border_style)
                        worksheet.write(ROW_DROP_CARD + 2, COL_DROP_CARD, 'Mod', self.row_title_style)
                        worksheet.write(ROW_DROP_CARD + 2, COL_DROP_CARD + 1, game_data.card_values[value[1]]['ref'], self.border_style)
                        worksheet.write(ROW_DROP_CARD + 3, COL_DROP_CARD, 'Rare mod', self.row_title_style)
                        worksheet.write(ROW_DROP_CARD + 3, COL_DROP_CARD + 1, game_data.card_values[value[2]]['ref'], self.border_style)
                    elif param_name in ['devour']:
                        worksheet.write(ROW_DEVOUR + 1, COL_DEVOUR, 'Low', self.row_title_style)
                        worksheet.write(ROW_DEVOUR + 1, COL_DEVOUR + 1, game_data.devour_values[value[0]]['ref'], self.border_style)
                        worksheet.write(ROW_DEVOUR + 2, COL_DEVOUR, 'Medium', self.row_title_style)
                        worksheet.write(ROW_DEVOUR + 2, COL_DEVOUR + 1, game_data.devour_values[value[1]]['ref'], self.border_style)
                        worksheet.write(ROW_DEVOUR + 3, COL_DEVOUR, 'High', self.row_title_style)
                        worksheet.write(ROW_DEVOUR + 3, COL_DEVOUR + 1, game_data.devour_values[value[2]]['ref'], self.border_style)
                    elif param_name in game_data.BYTE_FLAG_LIST:
                        for bit_name, bit_value in ennemy.info_stat_data[param_name].items():
                            worksheet.write(ROW_BYTE_FLAG + row_index['byte_flag'], COL_MISC, bit_name, self.row_title_style)
                            worksheet.write(ROW_BYTE_FLAG + row_index['byte_flag'], COL_MISC + 1, bit_value, self.border_style)
                            row_index['byte_flag'] += 1
                    elif param_name in ['renzokuken']:
                        row_index['renzokuken'] = 0
                        for el in value:
                            worksheet.write(ROW_RENZOKUKEN + row_index['renzokuken'] + 1, COL_MISC, 'Renzo value {}'.format(row_index['renzokuken'] + 1),
                                            self.row_title_style)
                            worksheet.write(ROW_RENZOKUKEN + row_index['renzokuken'] + 1, COL_MISC + 1, game_data.special_action[el]['ref'], self.border_style)
                            row_index['renzokuken'] += 1
                except IndexError as e:
                    raise IndexError("Unknown error on file {} for monster name {}: {}".format(key, ennemy.info_stat_data['monster_name'], e))
                # Looping on byte flag

            # Post validation
            self.__validation_post_process_all(worksheet, game_data)

            # Chart management
            chart_stat[ennemy].set_title({'name': 'Stat graph'})
            chart_stat[ennemy].set_x_axis({'name': 'Level'})
            chart_stat[ennemy].set_y_axis({'name': 'Stat'})
            chart_stat[ennemy].set_size({'width': STAT_GRAPH_WIDTH, 'height': STAT_GRAPH_HEIGHT})

            worksheet.insert_chart(STAT_GRAPH_CELL_PLACEMENT, chart_stat[ennemy])

            worksheet.autofit()

        # Creating reference data on last tab
        worksheet = self.workbook.add_worksheet(REF_DATA_SHEET_TITLE)
        worksheet.write_row(0, REF_DATA_COL_ABILITIES_TYPE, ['Ennemy type abilities', 'Ennemy abilities', 'Magic', 'Items', 'Devour', 'Card', 'Special action'],
                            cell_format=self.column_title_style)
        for index, el in game_data.ennemy_abilities_type_values.items():
            worksheet.write(index + 1, REF_DATA_COL_ABILITIES_TYPE, el['ref'], self.border_style)
        for index, el in game_data.ennemy_abilities_values.items():
            worksheet.write(index + 1, REF_DATA_COL_ABILITIES, el['ref'], self.border_style)
        for index, el in game_data.magic_values.items():
            worksheet.write(index + 1, REF_DATA_COL_MAGIC, el['ref'], self.border_style)
        for index, el in game_data.item_values.items():
            worksheet.write(index + 1, REF_DATA_COL_ITEM, el['ref'], self.border_style)
        for index, el in game_data.devour_values.items():
            worksheet.write(index + 1, REF_DATA_COL_DEVOUR, el['ref'], self.border_style)
        for index, el in game_data.card_values.items():
            worksheet.write(index + 1, REF_DATA_COL_CARD, el['ref'], self.border_style)
        for index, el in game_data.special_action.items():
            worksheet.write(index + 1, REF_DATA_COL_SPECIAL_ACTION, el['ref'], self.border_style)
        worksheet.autofit()

        # For test purpose
        temp_dict = {}
        for var_name, list_current_monst in game_data.game_info_test.items():
            if var_name == 'list_monster':
                pass
            new_key = var_name + "_sub"
            temp_list = game_data.game_info_test['list_monster'].copy()
            for temp in list_current_monst:
                if temp in temp_list:
                    temp_list.remove(temp)
            temp_dict[new_key] = temp_list.copy()
        game_data.game_info_test['sub'] = temp_dict

        #print(game_data.game_info_test)
        # End test purpose

        self.workbook.close()


class XlsxToDat():
    def __init__(self):
        pass

    def write_to_dat(self, ennemy_list, game_data: GameData, path: str):
        for key, ennemy in ennemy_list.items():
            ennemy.write_data_to_file(game_data, path)

    def import_from_xlsx(self, xlsx_file, ennemy, game_data, output_path, limit_file_index=-1):
        """
        As the module to write is different from the reading one, the one writing start at 0 for column and row, when this one, the reading, start at 1
        """
        wb = load_workbook(xlsx_file)
        for sheet in wb:

            if sheet.title == REF_DATA_SHEET_TITLE:
                continue

            ennemy_origin_file = sheet.cell(ROW_FILE_DATA + 1 + 1, COL_FILE_DATA + 1 + 1).value
            file_index = int(re.search(r'\d{3}', ennemy_origin_file).group())
            if limit_file_index >= 0 and limit_file_index != file_index:
                continue  # Just to not work on all files everytime
            ennemy[ennemy_origin_file] = Ennemy(game_data)
            current_ennemy = ennemy[ennemy_origin_file]
            current_ennemy.load_file_data(os.path.join(output_path, ennemy_origin_file), game_data)  # Loading the file to have all offset correct
            current_ennemy.analyse_loaded_data(game_data)
            current_ennemy.info_stat_data['monster_name'] = sheet.cell(ROW_MONSTER_NAME + 1, COL_MONSTER_INFO + 1 + 1).value
            current_ennemy.origin_file_name = ennemy_origin_file

            # Animation info on monster for the time beeing
            current_ennemy.model_animation_data['nb_animation'] = sheet.cell(ROW_MONSTER_NB_ANIMATION + 1, COL_MONSTER_INFO + 1 + 1).value

            # Stat reading
            row_index = 2
            for stat in game_data.stat_values:
                list_value = []
                for col_index in range(COL_STAT + 2, COL_STAT + 2 + 4):
                    list_value.append(sheet.cell(row_index, col_index).value)
                current_ennemy.info_stat_data[stat] = list_value
                row_index += 1

            # Def reading
            list_value = []
            for i in range(2, len(game_data.magic_type_values) + 1):
                list_value.append(sheet.cell(i, COL_DEF + 1 + 1).value)
            current_ennemy.info_stat_data['elem_def'] = list_value

            list_value = []
            for i in range(len(game_data.magic_type_values) + 1, len(game_data.magic_type_values) + len(game_data.status_values)):
                list_value.append(sheet.cell(i, COL_DEF + 1 + 1).value)
            current_ennemy.info_stat_data['status_def'] = list_value

            # Item read
            item = ['mag', 'mug', 'drop']
            sub_item = ['low_lvl', 'med_lvl', 'high_lvl']
            row_index = 2
            col_index = COL_ITEM + 1 + 1
            list_value = []
            for el in item:
                for sub in sub_item:
                    name = sub + "_" + el
                    for i in range(4):
                        id_value = int(sheet.cell(row_index + i, col_index).value.split(':')[0])
                        value = sheet.cell(row_index + i, col_index + 1).value
                        list_value.append({'ID': id_value, 'value': value})
                    current_ennemy.info_stat_data[name] = list_value
                    list_value = []

                    col_index += 2
                row_index += 4
                col_index = COL_ITEM + 1 + 1

            # Misc reading
            row_index = 2
            for misc in game_data.MISC_ORDER:
                value = sheet.cell(row_index, COL_MISC + 1 + 1).value
                if misc == "mug_rate" or misc == "drop_rate":
                    value = value * 100
                current_ennemy.info_stat_data[misc] = value
                row_index += 1

            # Abilities reading
            col_index = COL_ABILITIES + 1
            for abilities in game_data.ABILITIES_HIGHNESS_ORDER:
                ability_set = []
                for i in range(3, NB_MAX_ABILITIES + 3):
                    type = int(sheet.cell(i, col_index).value.split(':')[0])
                    ability_id = int(sheet.cell(i, col_index + 1).value.split(':')[0])
                    animation = int(sheet.cell(i, col_index + 2).value)
                    ability_set.append({'type': type, 'animation': animation, 'id': ability_id})
                current_ennemy.info_stat_data[abilities] = ability_set
                col_index += 3

            # Text reading
            combat_text_list = []
            for i in range(MAX_COMBAT_TXT):
                txt_value = sheet.cell(ROW_MONSTER_COMBAT_TEXT + 1 + i, 2).value
                if txt_value:
                    combat_text_list.append(txt_value)
                else:
                    break
            current_ennemy.info_stat_data['battle_text'] = combat_text_list

            # Card reading
            card_list = []
            for i in range(ROW_DROP_CARD + 1 + 1, ROW_DROP_CARD + 1 + 1 + len(current_ennemy.info_stat_data['card'])):
                cell_value = sheet.cell(i, COL_DROP_CARD + 1 + 1).value
                card_list.append(int(cell_value.split(':')[0]))
            current_ennemy.info_stat_data['card'] = card_list
            # Devour reading
            devour_list = []
            for i in range(ROW_DEVOUR + 1 + 1, ROW_DEVOUR + 1 + 1 + len(current_ennemy.info_stat_data['devour'])):
                cell_value = sheet.cell(i, COL_DEVOUR + 1 + 1).value
                devour_list.append(int(cell_value.split(':')[0]))
            current_ennemy.info_stat_data['devour'] = devour_list

            # Byte flag stat
            row_flag_index = 0
            for byte_flag_name in game_data.BYTE_FLAG_LIST:
                bit_list = []
                for i in range(8):
                    bit_list.append(sheet.cell(ROW_BYTE_FLAG + row_flag_index + 1, COL_MISC + 1 + 1).value)
                    row_flag_index += 1
                bit_list = bit_list[::-1]
                bits_string = ''.join(str(bit) for bit in bit_list)
                byte_result = int.from_bytes(bytes([int(bits_string, 2)]))
                current_ennemy.info_stat_data[byte_flag_name] = byte_result

            # Renzokuken reading
            renzokuken_list = []
            for i in range(ROW_RENZOKUKEN + 1 + 1, ROW_RENZOKUKEN + 1 + 1 + len(current_ennemy.info_stat_data['renzokuken'])):
                cell_value = sheet.cell(i, COL_MISC + 1 + 1).value
                renzokuken_list.append(int(cell_value.split(':')[0]))

            current_ennemy.info_stat_data['renzokuken'] = renzokuken_list
